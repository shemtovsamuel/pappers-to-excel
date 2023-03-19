#!/usr/bin/env python3

import requests
from openpyxl import Workbook, load_workbook
import os
from dotenv import load_dotenv
import datetime

load_dotenv()  # Charger les variables d'environnement du fichier .env

API_KEY = os.getenv("API_KEY")  # Récupérer la clé API depuis les variables d'environnement

def adjust_column_width(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Obtenir la lettre de la colonne
        for cell in col:
            if cell.coordinate in worksheet.merged_cells:  # Ignorer les cellules fusionnées
                continue
            try:  # Nécessaire pour éviter les erreurs sur les cellules vides
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Ajuster la largeur avec une marge supplémentaire
        worksheet.column_dimensions[column].width = adjusted_width

def get_value_safely(data, key, message):
    if data and isinstance(data, list) and len(data) > 0 and key in data[0]:
        return data[0][key]
    elif data and key in data:
        return data[key]
    else:
        print(message)
        return ""

def main(filename):
    siren = input("Entrez le numéro SIREN de l'entreprise à rechercher : ")
    url = f'https://api.pappers.fr/v2/entreprise?api_token={API_KEY}&siren={siren}'
    response = requests.get(url)

    if response.status_code == 200:
        data = response.json()
        nom_entreprise = get_value_safely(data, 'nom_entreprise', "Le nom de l'entreprise est vide")
        adresse_ligne_1 = get_value_safely(data.get('siege', {}), 'adresse_ligne_1', "L'adresse est vide")
        ville = get_value_safely(data.get('siege', {}), 'ville', "La ville est vide")
        code_postal = get_value_safely(data.get('siege', {}), 'code_postal', "Le code postal est vide")
        pays = get_value_safely(data.get('siege', {}), 'pays', "Le pays est vide")
        domaine_activite = get_value_safely(data, 'domaine_activite', "Le domaine d'activité est vide")
        dirigeant = get_value_safely(data.get('representants', [{}]), 'nom_complet', "Le nom du dirigeant est vide")
        date_naissance_dirigeant = get_value_safely(data.get('representants', [{}]), 'date_de_naissance_formate', "La date de naissance du dirigeant est vide")
        siret = get_value_safely(data.get('siege', {}), 'siret', "Le SIRET est vide")
        forme_juridique = get_value_safely(data, 'forme_juridique', "La forme juridique est vide")
        numero_tva_intracommunautaire = get_value_safely(data, 'numero_tva_intracommunautaire', "Le numéro de TVA intracommunautaire est vide")
        rcs = get_value_safely(data, 'numero_rcs', "Le numéro RCS est vide")
        capital_formate = get_value_safely(data, 'capital_formate', "Le capital social est vide")
        naf = get_value_safely(data.get('siege', {}), 'code_naf', "Le code NAF est vide")
        date_cloture = get_value_safely(data, 'prochaine_date_cloture_exercice_formate', "La date de clôture est vide")
        benefice_exercice_temp_nom = get_value_safely(data.get('beneficiaires_effectifs', [{}]), 'nom', "Le nom du bénéficiaire est vide")
        benefice_exercice_temp_prenom = get_value_safely(data.get('beneficiaires_effectifs', [{}]), 'prenom', "Le prénom du bénéficiaire est vide")
        benefice_exercice_nom = benefice_exercice_temp_prenom + " " + benefice_exercice_temp_nom

        # Charger le fichier Excel existant ou créer un nouveau
        if os.path.isfile(filename):
            wb = load_workbook(filename)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            # Écrire les en-têtes de colonnes
            ws.append(["Nom de la société", "Adresse", "Ville", "Code postal", "Pays", "Domaine d'activité", "Dirigeant", "Date de naissance du dirigeant", "Siret", "Forme juridique", "Numéro de TVA intracommunautaire", "RCS", "Capital social", "Code APE", "Date de clôture comptable", "Bénéficiaire"])

        # Écrire les informations de l'entreprise dans une ligne
        ws.append([nom_entreprise, adresse_ligne_1, ville, code_postal, pays, domaine_activite, dirigeant, date_naissance_dirigeant, siret, forme_juridique, numero_tva_intracommunautaire, rcs, capital_formate, naf, date_cloture, benefice_exercice_nom])

        # Ajuster la largeur des colonnes
        adjust_column_width(ws)

        # Sauvegarder le fichier Excel
        wb.save(filename)

        print(f"Les informations ont été ajoutées dans le fichier {filename}")
    else:
        if response.status_code == 400:
            print("Le numéro SIREN est invalide.")
        elif response.status_code == 403:
            print("Le nombre de requêtes autorisé a été dépassé.")
        else:
            print("Une erreur s'est produite lors de la récupération des données.")

if __name__ == "__main__":
    filename = input("Entrez le nom du fichier Excel (par exemple 'entreprises.xlsx') : ")
    if filename == "":
        filename = "entreprises.xlsx"
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    main(filename)
